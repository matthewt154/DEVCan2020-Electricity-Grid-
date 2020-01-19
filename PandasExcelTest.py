'''
from openpyxl import load_workbook
import pandas as pd
writer = pd.ExcelWriter('test.xlsx', engine='openpyxl') 
wb  = writer.book
df = pd.DataFrame({'Col_A': [1,2,3,4],
                  'Col_B': [5,6,7,8],
                  'Col_C': [0,0,0,0],
                  'Col_D': [13,14,15,16]})

df.to_excel(writer, index=False)
wb.save('test.xlsx')
'''
#Convert the list of lists to a dataframe and write to csv:

import urllib.request, urllib.error, urllib.parse

url = 'http://reports.ieso.ca/public/GenOutputbyFuelHourly/PUB_GenOutputbyFuelHourly_2020_v17.xml'

response = urllib.request.urlopen(url)
webContent = str(response.read())

f = open('Generator Output by Fuel Type Hourly Report.xml', 'w')
f.write(webContent)
f.close

#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
# Import BeautifulSoup
from bs4 import BeautifulSoup as bs


content = []
# Read the XML file
with open("Generator Output by Fuel Type Hourly Report.xml", "r") as file:
    # Read each line in the file, readlines() returns a list of lines
    content = file.readlines()
    # Combine the lines in the list into a string
    content = "".join(content)
    #print(content)
    bs_content = bs(content, "lxml")
result =bs_content.find_all("energyvalue")
#print (result)
newres = []
for all in result:
	newres.append(int(all.find("output").get_text()))
#print(newres)
finalOutput=[]
while (newres):
	sum=0
	for i in range(6):
		sum+=newres.pop() #add all 6 columns of energy types for final load 
	finalOutput.append(sum)
finalOutput.reverse()
print(finalOutput)

#add em all up 
import pandas as pd 
import openpyxl
from openpyxl.styles import Font 
#df=pd.DataFrame(columns=['Summary','Energy'])
wb =openpyxl.load_workbook('CANDev2020 Energy Loads.xlsx')
#writer =pd.ExcelWriter('CANDev2020 Energy Loads.xlsx')
'''
df1 = pd.read_excel('CANDev2020 Energy Loads.xlsx')
df2=pd.DataFrame()
df2.insert (0,'Hour','Energy')
df1.to_excel(writer,startrow=0,index=False)
df2.to_excel(writer,startrow=len(df1)+1,header=False,index=False)
'''
#wb=Workbook()
sheet=wb['testnew']
sheet.cell(row=1, column=2).value='Test'
#italic24Font = Font(size=24, italic=True) 
#sheet['A1']='italic24Font'
#df1.to_excel(writer,'testnew')
wb.save('CANDev2020 Energy Loads.xlsx')
#writer.save()
'''
i=408
for sublist in finalOutput:
	i+=1
	print(sublist)
	df=df.append({'Summary': i,'Energy':sublist}, ignore_index=True)
    
df.to_csv('AESO_Summary_.csv',index=False,encoding='cp1252')
'''

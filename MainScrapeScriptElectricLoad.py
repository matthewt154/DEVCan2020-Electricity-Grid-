import requests 
from bs4 import BeautifulSoup as bs
import bs4 
import os
import lxml.html as lh
import urllib.request, urllib.error, urllib.parse
import pandas as pd 
import openpyxl
import csv
#from apscheduler.schedulers 

#data sources 
#0- Ontario, 1-Alberta, 2-Nova Scotia, 3-New Brunswick
data_sources = ['http://ets.aeso.ca/ets_web/ip/Market/Reports/CSDReportServlet',
				'http://reports.ieso.ca/public/GenOutputbyFuelHourly/PUB_GenOutputbyFuelHourly_2020_v17.xml',
				'https://www.nspower.ca/oasis/monthly-reports/hourly-total-net-nova-scotia-load',
				'https://tso.nbpower.com/Public/en/system_information_archive.aspx']


def addToExcel (province,data):
	#method to add data to specific province worksheet 
	#param data is a list with indexes (0:date, 1:hour, 2:Load (MW))
	wb =openpyxl.load_workbook('CANDev2020 Energy Loads.xlsx') #excel book to write to 
	
	#sheet=wb[str(province)]
	sheet=wb[province]
	cnt=0; 
	while (True):
		cnt+=1 
		#print(cnt)
		if ((sheet.cell(row=cnt, column=1).value) is None):
			break
	#at empty cell, add new values 
	sheet.cell(row=cnt, column=1).value=data[0] #date
	sheet.cell(row=cnt, column=2).value=data[1] #hour
	sheet.cell(row=cnt, column=3).value=data[2] #load 
	wb.save('CANDev2020 Energy Loads.xlsx')


def scrapeAlberta(source,province):
	'''
	1-preliminary check to see if hour is correct (also record hour)
	2-Go to AIL value and isolate it
	3- Write it to correct spot in .csv file
	'''
	page = requests.get(source)
	tree = lh.fromstring(page.content)
	#get value of elements in td
	td_elements = tree.xpath('//td/text()')
	#storing the value in a list
	l_td = []
	for t in td_elements:
		l_td.append(t)
	#from our list, we get the value of the Alberta Internal Load
	for i in range(len(l_td) - 1):
		if l_td[i] == 'Alberta Internal Load (AIL)':
			alberta_value = l_td[i + 1]
	#Getting the time of last update for our value
	for i in range(len(l_td) - 1):
		if "Last Update" in l_td[i]:
			format_last_update = l_td[i]
	#Storing the data for the date of the last update of this information
	full_date_of_last_update = format_last_update.split()
	last_update_month = full_date_of_last_update[3].rstrip()
	last_update_date = full_date_of_last_update[4].rstrip()
	last_update_year = full_date_of_last_update[5].rstrip()
	last_update_time = full_date_of_last_update[6]
	#Checking to see if there is a comma in the string for the date so that we remove
	if "," in last_update_date:
		last_update_date = last_update_date.replace(",","")
	#print(alberta_value)
	real_date=str(last_update_date+"-"+last_update_month+'-'+last_update_year)
	#TO_DO put to excel file 
	data=[real_date, last_update_time,alberta_value]
	addToExcel(province, data)
	
	print ("Alberta scraped successfully")
	return 0

def scrapeOntario(source,province):
	'''
	1-Download correct XML file from source page (onto machine)
	2- Isolate correct date and time column from XML
	3- Get value (TOTAL OUTPUT) from last column
	4- Write to .csv file in correct column
	'''
	#downloading file from web to device and overwriting
	response = urllib.request.urlopen(source)
	webContent = str(response.read())
	filename='Generator Output by Fuel Type Hourly Report.xml'
	f = open(filename, 'w')
	f.write(webContent)
	f.close

	content = []
	# Read the XML file
	with open(filename, "r") as file:
	    # Read each line in the file, readlines() returns a list of lines
	    content = file.readlines()
	    # Combine the lines in the list into a string
	    content = "".join(content)
	    #print(content)
	    bs_content = bs(content, "lxml")
	result =bs_content.find_all("energyvalue")
	dateParse =bs_content.find_all("day")
	#print(dateParse)
	hourParse = bs_content.find_all("hour")
	#print (result)
	newres = []
	newDate = []
	newHour = []
	for all in result:
		newres.append(int(all.find("output").get_text()))
	#getting date 
	for all in dateParse:

		newDate.append(str(all)[5:15])

	#getting hour
	for all in hourParse:
		#print (str(all).find("/",7))
		newHour.append(str(all)[6:str(all).find("<",7)])
	#print(newHour)
	finalOutput=[]
	while (newres):
		sum=0
		for i in range(6):
			sum+=newres.pop() #add all 6 columns of energy types for final load 
		finalOutput.append(sum)
	#finalOutput.reverse() #most recent is first value 
	#getting date and hour 
	#print(newDate.pop())
	#print(newHour.pop())
	#print(finalOutput.pop())

	#TO_DO put to excel file 
	#updates every 24 hours so will pop 24 values 
	finalDate =newDate.pop()
	for i in range(24):
		finalHour= str(newHour.pop()+":00")
		data=[finalDate, finalHour, finalOutput.pop()]
		addToExcel(province,data)

	print ("Ontario scraped successfully")
	return 0

def scrapeNB(source):
	'''
	1- Change webpage to change date to last month then click Get Data 
	2- Navigates to .aspx file
	3. Go to first column to check date then second column for hour 
	4- 3rd column has load value we want. Write this to correct .csv file column 
	'''

	#TO_DO put to excel file 
	return 0

def scrapeNS(source):
	'''
	1- navigate to source page and open .htm file
	2- First column has date and hour. Split these so we can isolate date and time
	3- 2nd column has load value. Write this to correct .csv file 
	'''
	#Not complete code yet
	#receiving the url of the NS website
	page = requests.get(source)
	#saving data in a tree
	tree = lh.fromstring(page.content)
	#retrieving fref of csv file by parsing the website
	csv_files_link = tree.xpath('//a[@title = "CSV"]/@href')
	#the most recent csv file is in the first position
	most_recent_csv_file = csv_files_link[0]
	#opening csv file
	res = urllib.request.urlopen("https://www.nspower.ca"+most_recent_csv_file)
	#creating a list to save the data
	csv_lines = []
	#adding the rows of the csv file to the list
	for row in res:
                csv_lines.append(row)
        #grabbing the first 20 elements in csv_lines
        relevant_lines = csv_lines[:20]
        
        #unable to complete code but this what we have left so far. Issue: relevant_lines is read as a byte-code instead of a string code which means that we can't check if a string is in relevant_lines
        '''
        for i in range(len(relevant_lines)-1):
                if "Date/time,Load [MW]\r\n'" in relevant_lines[i]:
        '''             

	#TO_DO put to excel file 
	return 0

def mainRun():
	scrapeAlberta(data_sources[0],'Alberta')
	scrapeOntario(data_sources[1],'Ontario')
	return 0

#%%%%%%%%MAIN PROGRAM RUN TEST

mainRun() #running the main program


#%%%%%%%%%%%%%RUNNING PROGRAMfor time using scheduler
import datetime
import time
from apscheduler.scheduler import Scheduler

# Start the scheduler
sched = Scheduler()
sched.daemonic = False
sched.start()

def job_function():
    print("Collecting data..")
    mainRun()
    print(datetime.datetime.now())
    time.sleep(20)

# Schedules job_function to be run once (depends on when province website updates)
sched.add_cron_job(scrapeAlberta(data_sources[0],'Alberta'),  hour='0-23') #alberta updates every hour

sched.add_cron_job(scrapeOntario(data_sources[1],'Ontario'),  day='1-31') #Ontario updates every day 


